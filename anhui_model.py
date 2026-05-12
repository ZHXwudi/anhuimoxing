from __future__ import annotations

from dataclasses import dataclass, asdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


PRICE_PEAK = 1.3252
PRICE_HIGH = 1.0915
PRICE_FLAT = 0.6499
PRICE_LOW = 0.3128

BATTERY_RATIOS_BY_MODE = {
    1: [
        0.0,
        0.969299294671417,
        0.922118402184221,
        0.892376974402727,
        0.867145848671401,
        0.844843719695814,
        0.824999922362021,
        0.806251880135724,
        0.78866498842066,
        0.969299294671417,
        0.922118402184221,
        0.892376974402727,
        0.867145848671401,
        0.844843719695814,
        0.824999922362021,
        0.806251880135724,
        0.78866498842066,
    ],
    2: [
        0.0,
        0.96508443228735,
        0.911341951478372,
        0.877307270579312,
        0.848340721203993,
        0.822686149579489,
        0.799826112344998,
        0.778204067906047,
        0.757902914211625,
        0.96508443228735,
        0.911341951478372,
        0.877307270579312,
        0.848340721203993,
        0.822686149579489,
        0.799826112344998,
        0.778204067906047,
        0.757902914211625,
    ],
}


@dataclass(frozen=True)
class DeviceSpec:
    model: str
    mode: int
    power_kw: float
    rated_kwh: float
    actual_kwh: float
    device_cost_wan: float
    construction_cost_wan: float
    remark: str = ""


@dataclass(frozen=True)
class ModelParams:
    safety_factor: float = 0.85
    transformer_capacity_kva: float = 12550.0
    discount_rate: float = 0.8
    brokerage_rate: float = 0.13
    morning_limit_ratio: float = 1.0
    afternoon_limit_ratio: float = 1.0
    afternoon_charge_ratio: float = 0.0
    annual_decay_rate: float = 0.025
    system_efficiency_payback: float = 0.865
    operation_cost_rate: float = 0.01
    tax_rate: float = 0.05
    interest_rate: float = 0.04
    discharge_price_override: float = 1.0278588881151
    charge_price_override: float = 0.31275


@dataclass
class ConfigResult:
    rank: int
    model: str
    mode: int
    unit_count: int
    system_power_kw: float
    rated_kwh: float
    actual_kwh: float
    run_days: float
    payback_years: float | None
    total_discharge_kwh: float
    total_charge_kwh: float
    discharge_price: float
    charge_price: float
    price_spread: float
    first_year_income_wan: float
    initial_invest_wan: float
    final_cash_flow_wan: float
    balance_text: str
    score: float

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


def normalize_time(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    text = str(value).strip().replace("\t", "")
    if not text:
        return None
    for fmt in ("%H:%M:%S", "%H:%M", "%H:%M:%S.%f"):
        try:
            return datetime.strptime(text, fmt).strftime("%H:%M:%S")
        except ValueError:
            pass
    return None


def normalize_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def parse_minutes(value: str) -> int:
    h, m, s = (int(part) for part in value.split(":"))
    return h * 60 + m + s // 60


def time_in_range(value: str, start: str, end: str) -> bool:
    minute = parse_minutes(value)
    return parse_minutes(start) <= minute <= parse_minutes(end)


def load_device_specs(workbook_path: str | Path) -> list[DeviceSpec]:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb["设备规格"] if "设备规格" in wb.sheetnames else wb.worksheets[7]
    specs: list[DeviceSpec] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        mode = safe_float(row[1], -1)
        if mode not in (1, 2):
            continue
        specs.append(
            DeviceSpec(
                model=str(row[0]).strip(),
                mode=int(mode),
                power_kw=safe_float(row[2]),
                rated_kwh=safe_float(row[3]),
                actual_kwh=safe_float(row[4]),
                device_cost_wan=safe_float(row[5]),
                construction_cost_wan=safe_float(row[6]),
                remark=str(row[7] or ""),
            )
        )
    return specs


def load_pivot_from_workbook(workbook_path: str | Path) -> pd.DataFrame:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb["负荷-原始数据"] if "负荷-原始数据" in wb.sheetnames else wb.worksheets[0]
    records: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        day = normalize_date(values[1] if len(values) > 1 else None)
        slot = normalize_time(values[2] if len(values) > 2 else None)
        if day is None or slot is None:
            continue
        records.append({"日期": day, "时间": slot, "瞬时有功": safe_float(values[3])})
    raw = pd.DataFrame(records)
    if raw.empty:
        raise ValueError("工作簿里没有读取到有效负荷数据")
    return (
        raw.pivot_table(index="日期", columns="时间", values="瞬时有功", aggfunc="sum", fill_value=0.0)
        .reset_index()
        .sort_values("日期")
    )


def load_uploaded_pivot(file: Any) -> pd.DataFrame:
    name = getattr(file, "name", "")
    if name.lower().endswith((".xlsx", ".xlsm")):
        return load_pivot_from_workbook(file)
    raw = pd.read_csv(file)
    date_col = next((c for c in raw.columns if str(c).strip().lower() in ("日期", "date", "day")), raw.columns[0])
    time_col = next((c for c in raw.columns if str(c).strip().lower() in ("时间", "time", "时刻")), raw.columns[1])
    power_col = next(
        (c for c in raw.columns if str(c).strip().lower() in ("瞬时有功", "有功", "power", "load", "有功功率")),
        raw.columns[2],
    )
    raw = raw[[date_col, time_col, power_col]].copy()
    raw.columns = ["日期", "时间", "瞬时有功"]
    raw["日期"] = pd.to_datetime(raw["日期"]).dt.date
    raw["时间"] = raw["时间"].map(normalize_time)
    raw["瞬时有功"] = pd.to_numeric(raw["瞬时有功"], errors="coerce").fillna(0.0)
    return (
        raw.dropna(subset=["时间"])
        .pivot_table(index="日期", columns="时间", values="瞬时有功", aggfunc="sum", fill_value=0.0)
        .reset_index()
        .sort_values("日期")
    )


def build_power_matrix(
    pivot_df: pd.DataFrame,
    times: list[str],
    total_power: float,
    power_limit: float,
    afternoon_charge_ratio: float,
) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for _, row in pivot_df.iterrows():
        day = row["日期"]
        md = day.strftime("%m-%d")
        is_peak_date = ("01-01" <= md <= "01-31") or ("07-01" <= md <= "08-31")
        item: dict[str, Any] = {"日期": day}
        for slot in times:
            power = safe_float(row.get(slot))
            charge_val = max(0.0, min(total_power, power_limit - power))
            discharge_val = max(0.0, min(total_power, power - total_power * 0.1))
            if time_in_range(slot, "00:00:00", "06:45:00"):
                val = charge_val
            elif time_in_range(slot, "07:00:00", "10:45:00"):
                val = discharge_val
            elif time_in_range(slot, "11:00:00", "13:45:00"):
                val = charge_val
            elif time_in_range(slot, "14:00:00", "15:45:00"):
                val = min(charge_val, total_power * afternoon_charge_ratio)
            elif time_in_range(slot, "16:00:00", "17:45:00"):
                val = discharge_val if is_peak_date else discharge_val
            elif time_in_range(slot, "18:00:00", "22:45:00"):
                val = discharge_val
            elif time_in_range(slot, "23:00:00", "23:45:00"):
                val = charge_val
            else:
                val = 0.0
            item[slot] = val
        result.append(item)
    return result


def sum_power(row: dict[str, Any], slots: list[str]) -> float:
    return sum(safe_float(row.get(slot)) for slot in slots)


def calc_detailed_stats(
    new_df: list[dict[str, Any]],
    times: list[str],
    total_actual: float,
    morning_limit_ratio: float,
    afternoon_limit_ratio: float,
    afternoon_charge_ratio: float,
    charge_mode: int,
    daily_decay: float,
) -> list[dict[str, Any]]:
    low1_times = [t for t in times if time_in_range(t, "00:00:00", "06:45:00")]
    flat1_times = [t for t in times if time_in_range(t, "07:00:00", "10:45:00")]
    low2_times = [t for t in times if time_in_range(t, "11:00:00", "13:45:00")]
    flat2_times = [t for t in times if time_in_range(t, "14:00:00", "15:45:00")]
    high_short_times = [t for t in times if time_in_range(t, "16:00:00", "17:45:00")]
    high_extra_times = [t for t in times if time_in_range(t, "22:00:00", "22:45:00")]
    high_long_times = [t for t in times if time_in_range(t, "16:00:00", "22:45:00")]
    peak_times = [t for t in times if time_in_range(t, "18:00:00", "21:45:00")]

    daily_list: list[dict[str, Any]] = []
    prev_cycle2_remain = 0.0
    initial_capacity = total_actual
    for idx, row in enumerate(new_df):
        day = row["日期"]
        month = day.month
        md = day.strftime("%m-%d")
        current_cap = max(0.0, initial_capacity - daily_decay * initial_capacity * idx)
        morning_limit = current_cap * morning_limit_ratio
        afternoon_limit = current_cap * afternoon_limit_ratio

        theo_low1 = min(sum_power(row, low1_times) / 4, current_cap)
        theo_flat_charge1 = 0.0
        theo_peak1 = 0.0
        theo_high1 = 0.0
        theo_flat_discharge1 = min(sum_power(row, flat1_times) / 4, current_cap)
        theo_low2 = min(sum_power(row, low2_times) / 4, current_cap)
        theo_flat_charge2 = min((sum_power(row, flat2_times) / 4) * afternoon_charge_ratio, current_cap)
        if ("01-01" <= md <= "01-31") or ("07-01" <= md <= "08-31") or ("12-01" <= md <= "12-31"):
            theo_peak2 = min(sum_power(row, peak_times) / 4, current_cap)
        else:
            theo_peak2 = 0.0
        high_sum = sum_power(row, high_short_times) + sum_power(row, high_extra_times) if month in [1, 7, 8, 12] else sum_power(row, high_long_times)
        theo_high2 = min(high_sum / 4, current_cap)

        actual_low1 = min(theo_low1, current_cap - prev_cycle2_remain) if idx else min(theo_low1, current_cap)
        actual_low1 = max(0.0, actual_low1)
        avail_flat_charge1 = max(0.0, current_cap - actual_low1 - (prev_cycle2_remain if idx else 0.0))
        actual_flat_charge1 = min(theo_flat_charge1, avail_flat_charge1)
        total_charge1 = actual_low1 + actual_flat_charge1 + (prev_cycle2_remain if idx else 0.0)
        discharge_limit1 = max(0.0, min(total_charge1, morning_limit))
        actual_high1 = min(discharge_limit1, theo_high1)
        remain1 = discharge_limit1 - actual_high1
        actual_peak1 = min(remain1, theo_peak1)
        remain1 -= actual_peak1
        actual_flat_discharge1 = min(remain1, theo_flat_discharge1)
        total_discharge1 = actual_high1 + actual_peak1 + actual_flat_discharge1
        cycle1_remain = total_charge1 - total_discharge1

        avail_low2 = max(0.0, current_cap - cycle1_remain)
        actual_low2 = min(theo_low2, avail_low2)
        actual_flat_charge2 = min(theo_flat_charge2, max(0.0, avail_low2 - actual_low2))
        total_charge2 = cycle1_remain + actual_low2 + actual_flat_charge2
        discharge_limit2 = max(0.0, min(total_charge2, afternoon_limit))
        actual_peak2 = min(discharge_limit2, theo_peak2)
        remain2 = discharge_limit2 - actual_peak2
        actual_high2 = min(remain2, theo_high2)
        total_discharge2 = actual_peak2 + actual_high2
        cycle2_remain = total_charge2 - total_discharge2
        prev_cycle2_remain = cycle2_remain

        cycle1 = actual_peak1 + actual_high1 + actual_flat_discharge1
        cycle2 = actual_peak2 + actual_high2
        total_discharge = cycle1 + cycle2
        eq_days = total_discharge / current_cap / 2 if charge_mode == 2 and current_cap else total_discharge / current_cap if current_cap else 0.0
        full_days = total_discharge / (initial_capacity * 2) if initial_capacity else 0.0
        low_charge = actual_low1 + actual_low2
        flat_charge = actual_flat_charge1 + actual_flat_charge2
        peak_discharge = actual_peak1 + actual_peak2
        high_discharge = actual_high1 + actual_high2
        flat_discharge = actual_flat_discharge1

        daily_list.append(
            {
                "日期": day,
                "月份": month,
                "折算天数": eq_days,
                "满容量天数": full_days,
                "低谷充电": low_charge,
                "平段充电": flat_charge,
                "尖峰放电": peak_discharge,
                "高峰放电": high_discharge,
                "平段放电": flat_discharge,
                "总放电": total_discharge,
                "理论总可充电量": min(theo_low1 + theo_flat_charge1 + theo_low2 + theo_flat_charge2, current_cap * 2),
                "理论总可放电量": min(theo_peak1 + theo_high1 + theo_flat_discharge1 + theo_peak2 + theo_high2, current_cap * 2),
                "充电容量": current_cap,
            }
        )
    return daily_list


def aggregate_monthly(daily_list: list[dict[str, Any]]) -> list[dict[str, Any]]:
    month_days = {1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31}
    month_type = {1: "特殊", 2: "正常", 3: "正常", 4: "正常", 5: "正常", 6: "正常", 7: "特殊", 8: "特殊", 9: "正常", 10: "正常", 11: "正常", 12: "特殊"}
    grouped: dict[int, dict[str, float]] = {}
    for item in daily_list:
        month = int(item["月份"])
        dest = grouped.setdefault(
            month,
            {
                "天数": 0.0,
                "折算天数": 0.0,
                "谷充电量": 0.0,
                "平充电量": 0.0,
                "尖放电量": 0.0,
                "峰放电量": 0.0,
                "平放电量": 0.0,
                "总放电量": 0.0,
                "总理论可充电": 0.0,
                "总理论可放电": 0.0,
                "总充电容量2倍": 0.0,
            },
        )
        dest["天数"] += item["满容量天数"]
        dest["折算天数"] += item["折算天数"]
        dest["谷充电量"] += item["低谷充电"]
        dest["平充电量"] += item["平段充电"]
        dest["尖放电量"] += item["尖峰放电"]
        dest["峰放电量"] += item["高峰放电"]
        dest["平放电量"] += item["平段放电"]
        dest["总放电量"] += item["总放电"]
        dest["总理论可充电"] += item["理论总可充电量"]
        dest["总理论可放电"] += item["理论总可放电量"]
        dest["总充电容量2倍"] += item["充电容量"] * 2

    rows: list[dict[str, Any]] = []
    for month in range(1, 13):
        d = grouped.get(month)
        if not d:
            rows.append({"月份": month, "分类": month_type[month], "天数": 0.0, "折算天数": 0.0, "综合充放电文本": "无数据"})
            continue
        total_charge = d["谷充电量"] + d["平充电量"]
        total_discharge = d["总放电量"]
        ratio_charge = d["总理论可充电"] / d["总充电容量2倍"] if d["总充电容量2倍"] else 0.0
        ratio_discharge = d["总理论可放电"] / d["总充电容量2倍"] if d["总充电容量2倍"] else 0.0
        diff = ratio_charge - ratio_discharge
        text = "消纳不足" if diff >= 0.05 else "充电不足" if diff <= -0.05 else "充放平衡"
        rows.append(
            {
                "月份": month,
                "分类": month_type[month],
                "天数": d["天数"],
                "折算天数": d["折算天数"],
                "谷充电量(度)": d["谷充电量"],
                "平充电量(度)": d["平充电量"],
                "总充电量(度)": total_charge,
                "尖放电量(度)": d["尖放电量"],
                "峰放电量(度)": d["峰放电量"],
                "平放电量(度)": d["平放电量"],
                "总放电量(度)": total_discharge,
                "谷充电占比": d["谷充电量"] / total_charge if total_charge else 0.0,
                "平充电占比": d["平充电量"] / total_charge if total_charge else 0.0,
                "尖放电占比": d["尖放电量"] / total_discharge if total_discharge else 0.0,
                "峰放电占比": d["峰放电量"] / total_discharge if total_discharge else 0.0,
                "平放电占比": d["平放电量"] / total_discharge if total_discharge else 0.0,
                "每月天数": month_days[month],
                "理论可充电情况": ratio_charge,
                "理论可放电情况": ratio_discharge,
                "综合充放电数值": min(ratio_charge, ratio_discharge),
                "综合充放电文本": text,
            }
        )

    total_keys = ["天数", "折算天数", "谷充电量(度)", "平充电量(度)", "总充电量(度)", "尖放电量(度)", "峰放电量(度)", "平放电量(度)", "总放电量(度)"]
    total = {key: sum(safe_float(row.get(key)) for row in rows) for key in total_keys}
    total_charge = total["总充电量(度)"]
    total_discharge = total["总放电量(度)"]
    total_charge_condition = sum(safe_float(row.get("理论可充电情况")) for row in rows) / 12
    total_discharge_condition = sum(safe_float(row.get("理论可放电情况")) for row in rows) / 12
    rows.append(
        {
            "月份": "总计",
            "分类": "",
            **total,
            "谷充电占比": total["谷充电量(度)"] / total_charge if total_charge else 0.0,
            "平充电占比": total["平充电量(度)"] / total_charge if total_charge else 0.0,
            "尖放电占比": total["尖放电量(度)"] / total_discharge if total_discharge else 0.0,
            "峰放电占比": total["峰放电量(度)"] / total_discharge if total_discharge else 0.0,
            "平放电占比": total["平放电量(度)"] / total_discharge if total_discharge else 0.0,
            "每月天数": 365,
            "理论可充电情况": total_charge_condition,
            "理论可放电情况": total_discharge_condition,
            "综合充放电数值": min(total_charge_condition, total_discharge_condition),
            "综合充放电文本": "",
        }
    )
    return rows


def weighted_prices(total_row: dict[str, Any]) -> tuple[float, float]:
    discharge_price = (
        PRICE_PEAK * safe_float(total_row.get("尖放电占比"))
        + PRICE_HIGH * safe_float(total_row.get("峰放电占比"))
        + PRICE_FLAT * safe_float(total_row.get("平放电占比"))
    )
    charge_price = PRICE_LOW * safe_float(total_row.get("谷充电占比")) + PRICE_FLAT * safe_float(total_row.get("平充电占比"))
    return discharge_price, charge_price


def _slots_between(times: list[str], start: str, end: str) -> list[str]:
    return [slot for slot in times if start <= slot <= end]


MORNING_FIXED_FLAT = ["09:00:00", "09:15:00", "09:30:00", "09:45:00", "10:00:00", "10:15:00", "10:30:00", "10:45:00"]
OLD_RULE_1 = ["15:00:00", "15:15:00", "15:30:00", "15:45:00"]
OLD_RULE_2 = [
    "16:00:00", "16:15:00", "16:30:00", "16:45:00", "17:00:00", "17:15:00", "17:30:00", "17:45:00",
    "18:00:00", "18:15:00", "18:30:00", "18:45:00", "19:00:00", "19:15:00", "19:30:00", "19:45:00",
    "20:00:00", "20:15:00", "20:30:00", "20:45:00", "21:00:00", "21:15:00", "21:30:00", "21:45:00",
    "22:00:00", "22:15:00", "22:30:00", "22:45:00",
]
OLD_RULE_3 = OLD_RULE_2[:16]
OLD_RULE_4 = ["15:00:00", "15:15:00", "15:30:00", "15:45:00"] + OLD_RULE_3[:12]
NEW_RULE_1 = OLD_RULE_1
NEW_RULE_2 = OLD_RULE_3
NEW_RULE_3 = OLD_RULE_2[16:]
NEW_RULE_23 = ["23:00:00", "23:15:00", "23:30:00", "23:45:00"]
DEC_LATE_19 = ["19:00:00", "19:15:00", "19:30:00", "19:45:00"]


def build_power_matrix_excel_rules(
    pivot_df: pd.DataFrame,
    times: list[str],
    total_power: float,
    power_limit: float,
) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for _, row in pivot_df.iterrows():
        day = row["日期"]
        md = day.strftime("%m-%d")
        month = day.month
        item: dict[str, Any] = {"日期": day}
        for idx, slot in enumerate(times):
            power = safe_float(row.get(slot))
            charge_val = max(0.0, min(total_power, power_limit - power))
            discharge_val = max(0.0, min(total_power, power - total_power * 0.1))

            if slot in DEC_LATE_19 and md >= "12-16":
                val = discharge_val
            elif slot in NEW_RULE_23:
                if ("01-01" <= md <= "06-30") or ("10-01" <= md <= "12-31"):
                    val = charge_val
                elif "07-01" <= md <= "09-30":
                    val = discharge_val
                else:
                    val = 0.0
            elif slot in NEW_RULE_1 and "02-01" <= md <= "12-15":
                val = charge_val
            elif slot in NEW_RULE_2 and "09-01" <= md <= "12-15":
                val = discharge_val
            elif slot in NEW_RULE_3 and md >= "07-15":
                val = discharge_val
            elif slot in OLD_RULE_4 and md >= "12-16":
                val = discharge_val
            elif slot in OLD_RULE_3 and "07-15" <= md <= "08-31":
                val = discharge_val
            elif slot in OLD_RULE_2 and md <= "07-14":
                val = discharge_val
            elif slot in OLD_RULE_1 and md <= "01-31":
                val = discharge_val
            elif slot in MORNING_FIXED_FLAT:
                val = discharge_val
            elif idx < 24:
                val = charge_val
            elif 24 <= idx <= 35:
                val = charge_val if 7 <= month <= 9 else discharge_val
            elif 44 <= idx <= 47:
                val = charge_val if 2 <= month <= 11 else discharge_val
            elif 48 <= idx <= 59:
                val = charge_val
            else:
                val = 0.0
            item[slot] = val
        result.append(item)
    return result


def calc_detailed_stats_excel_rules(
    new_df: list[dict[str, Any]],
    times: list[str],
    total_actual: float,
    morning_limit_ratio: float,
    afternoon_limit_ratio: float,
    afternoon_charge_ratio: float,
    charge_mode: int,
    daily_decay: float,
) -> list[dict[str, Any]]:
    low1 = _slots_between(times, "00:00:00", "05:45:00")
    high1 = _slots_between(times, "06:00:00", "07:45:00")
    flat_1_12 = _slots_between(times, "06:00:00", "11:45:00")
    flat_2_6_10_11 = _slots_between(times, "08:00:00", "10:45:00")
    flat_7_9 = _slots_between(times, "09:00:00", "10:45:00")
    low2_1_12 = _slots_between(times, "12:00:00", "13:45:00")
    low2_2_6_10_11 = _slots_between(times, "11:00:00", "13:45:00")
    low2_7_9 = _slots_between(times, "11:00:00", "12:45:00")
    flat2_1_12 = _slots_between(times, "14:00:00", "14:45:00")
    flat2_2_6_10_11 = _slots_between(times, "14:00:00", "15:45:00")
    flat2_7_9 = _slots_between(times, "13:00:00", "15:45:00")
    peak2_1_dec = _slots_between(times, "19:00:00", "20:45:00")
    peak2_7_8 = _slots_between(times, "20:00:00", "21:45:00")
    high2_1_part1 = _slots_between(times, "15:00:00", "18:45:00")
    high2_1_part2 = _slots_between(times, "21:00:00", "22:45:00")
    high2_2_714 = _slots_between(times, "16:00:00", "21:45:00")
    high2_78_part1 = _slots_between(times, "16:00:00", "19:45:00")
    high2_78_part2 = _slots_between(times, "22:00:00", "23:45:00")
    high2_9 = _slots_between(times, "16:00:00", "23:45:00")
    high2_10_11 = _slots_between(times, "16:00:00", "21:45:00")
    high2_dec_1_15 = _slots_between(times, "15:00:00", "22:45:00")

    daily: list[dict[str, Any]] = []
    prev_cycle2_remain = 0.0
    initial_capacity = total_actual
    for idx, row in enumerate(new_df):
        day = row["日期"]
        month = day.month
        md = day.strftime("%m-%d")
        cap = max(0.0, initial_capacity - daily_decay * initial_capacity * idx)
        morning_limit = cap * morning_limit_ratio
        afternoon_limit = cap * afternoon_limit_ratio

        theo_low1 = min(sum_power(row, low1) / 4, cap)
        theo_flat_charge1 = 0.0
        theo_peak1 = 0.0
        theo_high1 = min(sum_power(row, high1) / 4, cap) if (("02-01" <= md <= "06-30") or ("10-01" <= md <= "11-30")) else 0.0
        if ("01-01" <= md <= "01-31") or ("12-01" <= md <= "12-31"):
            flat1_cols = flat_1_12
        elif ("02-01" <= md <= "06-30") or ("10-01" <= md <= "11-30"):
            flat1_cols = flat_2_6_10_11
        elif "07-01" <= md <= "09-30":
            flat1_cols = flat_7_9
        else:
            flat1_cols = []
        theo_flat_discharge1 = min(sum_power(row, flat1_cols) / 4, cap) if flat1_cols else 0.0

        if ("01-01" <= md <= "01-31") or ("12-01" <= md <= "12-31"):
            low2_cols = low2_1_12
            flat2_cols = flat2_1_12
        elif ("02-01" <= md <= "06-30") or ("10-01" <= md <= "11-30"):
            low2_cols = low2_2_6_10_11
            flat2_cols = flat2_2_6_10_11
        elif "07-01" <= md <= "09-30":
            low2_cols = low2_7_9
            flat2_cols = flat2_7_9
        else:
            low2_cols = []
            flat2_cols = []
        theo_low2 = min(sum_power(row, low2_cols) / 4, cap) if low2_cols else 0.0
        theo_flat_charge2 = min((sum_power(row, flat2_cols) / 4) * afternoon_charge_ratio, cap) if flat2_cols else 0.0

        if "01-01" <= md <= "01-31":
            theo_peak2 = min(sum_power(row, peak2_1_dec) / 4, cap)
        elif "02-01" <= md <= "07-14":
            theo_peak2 = 0.0
        elif "07-15" <= md <= "08-31":
            theo_peak2 = min(sum_power(row, peak2_7_8) / 4, cap)
        elif "09-01" <= md <= "12-15":
            theo_peak2 = 0.0
        elif "12-16" <= md <= "12-31":
            theo_peak2 = min(sum_power(row, peak2_1_dec) / 4, cap)
        else:
            theo_peak2 = 0.0

        if "01-01" <= md <= "01-31":
            theo_high2 = min((sum_power(row, high2_1_part1) + sum_power(row, high2_1_part2)) / 4, cap)
        elif "02-01" <= md <= "07-14":
            theo_high2 = min(sum_power(row, high2_2_714) / 4, cap)
        elif "07-15" <= md <= "08-31":
            theo_high2 = min((sum_power(row, high2_78_part1) + sum_power(row, high2_78_part2)) / 4, cap)
        elif "09-01" <= md <= "09-30":
            theo_high2 = min(sum_power(row, high2_9) / 4, cap)
        elif "10-01" <= md <= "11-30":
            theo_high2 = min(sum_power(row, high2_10_11) / 4, cap)
        elif "12-01" <= md <= "12-15":
            theo_high2 = min(sum_power(row, high2_dec_1_15) / 4, cap)
        elif "12-16" <= md <= "12-31":
            theo_high2 = min((sum_power(row, high2_1_part1) + sum_power(row, high2_1_part2)) / 4, cap)
        else:
            theo_high2 = 0.0
        theo_flat_discharge2 = 0.0

        actual_low1 = min(theo_low1, cap - prev_cycle2_remain) if idx else min(theo_low1, cap)
        actual_low1 = max(0.0, actual_low1)
        actual_flat_charge1 = min(theo_flat_charge1, max(0.0, cap - actual_low1 - (prev_cycle2_remain if idx else 0.0)))
        total_charge1 = actual_low1 + actual_flat_charge1 + (prev_cycle2_remain if idx else 0.0)
        discharge_limit1 = max(0.0, min(total_charge1, morning_limit))
        actual_high1 = min(discharge_limit1, theo_high1)
        remain1 = discharge_limit1 - actual_high1
        actual_peak1 = min(remain1, theo_peak1)
        remain1 -= actual_peak1
        actual_flat_discharge1 = min(remain1, theo_flat_discharge1)
        cycle1_remain = total_charge1 - actual_high1 - actual_peak1 - actual_flat_discharge1

        actual_low2 = min(theo_low2, max(0.0, cap - cycle1_remain))
        actual_flat_charge2 = min(theo_flat_charge2, max(0.0, cap - cycle1_remain - actual_low2))
        total_charge2 = cycle1_remain + actual_low2 + actual_flat_charge2
        discharge_limit2 = max(0.0, min(total_charge2, afternoon_limit))
        actual_peak2 = min(discharge_limit2, theo_peak2)
        remain2 = discharge_limit2 - actual_peak2
        actual_high2 = min(remain2, theo_high2)
        remain2 -= actual_high2
        actual_flat_discharge2 = min(remain2, theo_flat_discharge2)
        cycle2_remain = total_charge2 - actual_peak2 - actual_high2 - actual_flat_discharge2
        prev_cycle2_remain = cycle2_remain

        cycle1 = actual_peak1 + actual_high1 + actual_flat_discharge1
        cycle2 = actual_peak2 + actual_high2 + actual_flat_discharge2
        total_discharge = cycle1 + cycle2
        daily.append(
            {
                "日期": day,
                "月份": month,
                "折算天数": total_discharge / cap / 2 if charge_mode == 2 and cap else total_discharge / cap if cap else 0.0,
                "满容量天数": total_discharge / initial_capacity / 2 if initial_capacity else 0.0,
                "低谷充电": actual_low1 + actual_low2,
                "平段充电": actual_flat_charge1 + actual_flat_charge2,
                "尖峰放电": actual_peak1 + actual_peak2,
                "高峰放电": actual_high1 + actual_high2,
                "平段放电": actual_flat_discharge1 + actual_flat_discharge2,
                "总放电": total_discharge,
                "理论总可充电量": min(theo_low1 + theo_flat_charge1 + theo_low2 + theo_flat_charge2, cap * 2),
                "理论总可放电量": min(theo_peak1 + theo_high1 + theo_flat_discharge1 + theo_peak2 + theo_high2 + theo_flat_discharge2, cap * 2),
                "充电容量": cap,
            }
        )
    return daily


def payback_summary(
    run_days: float,
    discharge_tax: float,
    charge_tax: float,
    total_rated: float,
    total_actual: float,
    unit_count: int,
    charge_mode: int,
    device_cost: float,
    construction_cost: float,
    params: ModelParams,
) -> dict[str, Any]:
    total_device_cost = device_cost * unit_count
    total_construction_cost = construction_cost * unit_count
    invest_yr9 = 0.4 * total_rated / 10 / 1.13
    brokerage_yr0 = total_rated * params.brokerage_rate / 10
    invest_tax_yr1 = total_device_cost + total_construction_cost + brokerage_yr0
    invest_no_tax_yr1 = total_device_cost / 1.13 + total_construction_cost + brokerage_yr0
    input_tax = invest_tax_yr1 - invest_no_tax_yr1
    battery_ratio = BATTERY_RATIOS_BY_MODE.get(charge_mode, BATTERY_RATIOS_BY_MODE[2])
    cycle_multiplier = 2 if charge_mode == 2 else 1

    power_station_invest = [0.0] * 17
    power_station_invest[0] = total_device_cost
    power_station_invest[8] = invest_yr9 * 1.13
    construction_invest = [0.0] * 17
    construction_invest[0] = total_construction_cost
    brokerage = [0.0] * 17
    brokerage[0] = brokerage_yr0

    charge_kwh = [0.0] * 17
    discharge_kwh = [0.0] * 17
    cash_in = [0.0] * 17
    for n in range(1, 17):
        charge_kwh[n] = (run_days * total_actual) * cycle_multiplier / 10000 / (1 - (1 - params.system_efficiency_payback) / 2) * battery_ratio[n]
        discharge_kwh[n] = params.system_efficiency_payback * charge_kwh[n]
        cash_in[n] = (discharge_kwh[n] * discharge_tax - charge_kwh[n] * charge_tax) * params.discount_rate

    operation_cost = [0.0] * 17
    insurance_cost = [0.0] * 17
    output_tax = [0.0] * 17
    for n in range(1, 17):
        operation_cost[n] = params.operation_cost_rate * 20 * unit_count
        insurance_cost[n] = invest_no_tax_yr1 * 0.0015
        output_tax[n] = cash_in[n] - cash_in[n] / 1.13

    vat_tax = [0.0] * 17
    vat_tax[0] = -input_tax
    for n in range(1, 17):
        vat_tax[n] = output_tax[n]
    pay_vat = [0.0] * 17
    for n in range(1, 17):
        sum_vat = sum(vat_tax[: n + 1])
        paid_before = sum(pay_vat[:n])
        if sum_vat > 0 and paid_before == 0:
            pay_vat[n] = sum_vat
        elif sum_vat > 0 and paid_before > 0:
            pay_vat[n] = vat_tax[n]

    tax_surcharge = [0.0] * 17
    depreciation = [0.0] * 17
    depr_yr1_8 = invest_no_tax_yr1 * 0.95 / 8
    depr_yr9_16 = power_station_invest[8] * 0.95 / (1.13 * 8)
    for n in range(1, 17):
        tax_surcharge[n] = cash_in[n] * 3 / 10000 + pay_vat[n] * 0.12
        depreciation[n] = depr_yr1_8 if n <= 8 else depr_yr9_16

    cash_out = [0.0] * 17
    net_cash_flow = [0.0] * 17
    cum_cash_flow = [0.0] * 17
    interest = [0.0] * 17
    income_tax = [0.0] * 17
    project_pre_tax_profit = [0.0] * 17
    judge_vat = [0.0] * 17
    indirect_cum_cash_flow = [0.0] * 17

    cash_out[0] = total_device_cost + total_construction_cost + brokerage[0]
    net_cash_flow[0] = -cash_out[0]
    cum_cash_flow[0] = net_cash_flow[0]
    indirect_cum_cash_flow[0] = -invest_no_tax_yr1 + vat_tax[0]

    for n in range(1, 17):
        interest[n] = abs(cum_cash_flow[n - 1]) * params.interest_rate if cum_cash_flow[n - 1] < 0 else 0.0
        taxable_base = cash_in[n] - operation_cost[n] - insurance_cost[n] - tax_surcharge[n] - interest[n] - depreciation[n] - output_tax[n]
        income_tax[n] = taxable_base * params.tax_rate
        cash_out[n] = power_station_invest[n] + brokerage[n] + operation_cost[n] + insurance_cost[n] + tax_surcharge[n] + interest[n] + income_tax[n] + pay_vat[n]
        net_cash_flow[n] = cash_in[n] - cash_out[n]
        cum_cash_flow[n] = cum_cash_flow[n - 1] + net_cash_flow[n]
        project_pre_tax_profit[n] = round(taxable_base, 2)
        judge_vat[n] = pay_vat[n] if pay_vat[n] > 0 else 0.0
        indirect_cum_cash_flow[n] = project_pre_tax_profit[n] - income_tax[n] + depreciation[n] + indirect_cum_cash_flow[n - 1] + output_tax[n] - judge_vat[n]

    payback = None
    for idx, value in enumerate(indirect_cum_cash_flow):
        if value >= 0:
            if idx == 0:
                payback = 0.0
            else:
                prev = indirect_cum_cash_flow[idx - 1]
                denom = value - prev
                payback = round((idx - 1) + abs(prev) / denom, 2) if denom else float(idx)
            break

    payback_yearly: list[str] = []
    for idx, value in enumerate(indirect_cum_cash_flow):
        if payback is None:
            payback_yearly.append("未回收")
        elif value >= 0:
            payback_yearly.append(f"{payback:.2f}年（已回收）")
        elif idx < 16:
            next_value = indirect_cum_cash_flow[idx + 1]
            denom = next_value - value
            payback_yearly.append(f"{abs(value) / denom:.2f}年（预测）" if denom else "未回收")
        else:
            payback_yearly.append("未回收")

    columns = ["项目", *[f"第{n}年" for n in range(17)], "合计"]
    table_rows = [
        ["电池容量", *[f"{value * 100:.6f}%" if value else "0.000000%" for value in battery_ratio], "-"],
        ["充电量（万度）", *charge_kwh, round(sum(charge_kwh), 4)],
        ["放电量（万度）", *discharge_kwh, round(sum(discharge_kwh), 4)],
        ["现金流入（电费收入）", *cash_in, round(sum(cash_in), 4)],
        ["电费收入", *cash_in, round(sum(cash_in), 4)],
        ["现金流出", *cash_out, round(sum(cash_out), 4)],
        ["电站投资", *power_station_invest, round(sum(power_station_invest), 4)],
        ["施工", *construction_invest, round(sum(construction_invest), 4)],
        ["居间", *brokerage, round(sum(brokerage), 4)],
        ["运营成本", *operation_cost, round(sum(operation_cost), 4)],
        ["保险费用", *insurance_cost, round(sum(insurance_cost), 4)],
        ["税金及附加", *tax_surcharge, round(sum(tax_surcharge), 4)],
        ["利息", *interest, round(sum(interest), 4)],
        ["税费（所得税，非绝对值）", *income_tax, round(sum(income_tax), 4)],
        ["支付增值税", *pay_vat, round(sum(pay_vat), 4)],
        ["现金净流入", *net_cash_flow, round(sum(net_cash_flow), 4)],
        ["累计现金流量", *cum_cash_flow, round(cum_cash_flow[-1], 4)],
        ["项目折旧", *depreciation, round(sum(depreciation), 4)],
        ["增值税", *vat_tax, round(sum(vat_tax), 4)],
        ["销项税", *output_tax, round(sum(output_tax), 4)],
        ["项目税前利润", *project_pre_tax_profit, round(sum(project_pre_tax_profit), 2)],
        ["判断", *judge_vat, round(sum(judge_vat), 4)],
        ["间接法累计现金流", *indirect_cum_cash_flow, round(indirect_cum_cash_flow[-1], 4)],
        ["静态投资回收期", *payback_yearly, f"{payback:.2f}年（已回收）" if payback is not None else "全周期无法回收投资"],
    ]

    return {
        "payback_years": payback,
        "first_year_income_wan": cash_in[1],
        "initial_invest_wan": cash_out[0],
        "final_cash_flow_wan": indirect_cum_cash_flow[-1],
        "battery_ratio": battery_ratio,
        "charge_kwh_10k": charge_kwh,
        "discharge_kwh_10k": discharge_kwh,
        "cash_in_wan": cash_in,
        "cash_out_wan": cash_out,
        "net_cash_flow_wan": net_cash_flow,
        "cum_cash_flow_wan": cum_cash_flow,
        "indirect_cum_cash_flow_wan": indirect_cum_cash_flow,
        "payback_table": pd.DataFrame(table_rows, columns=columns),
    }


def evaluate_config(pivot_df: pd.DataFrame, spec: DeviceSpec, unit_count: int, params: ModelParams) -> tuple[ConfigResult, pd.DataFrame, dict[str, Any]]:
    times = sorted([col for col in pivot_df.columns if col != "日期" and normalize_time(col) is not None], key=parse_minutes)
    total_power = spec.power_kw * unit_count
    total_rated = spec.rated_kwh * unit_count
    total_actual = spec.actual_kwh * unit_count
    power_limit = params.safety_factor * params.transformer_capacity_kva
    daily_decay = params.annual_decay_rate / 365
    power_matrix = build_power_matrix_excel_rules(pivot_df, times, total_power, power_limit)
    daily = calc_detailed_stats_excel_rules(
        power_matrix,
        times,
        total_actual,
        params.morning_limit_ratio,
        params.afternoon_limit_ratio,
        params.afternoon_charge_ratio,
        spec.mode,
        daily_decay,
    )
    monthly = aggregate_monthly(daily)
    monthly_df = pd.DataFrame(monthly)
    total_row = monthly[-1]
    discharge_price = params.discharge_price_override if params.discharge_price_override > 0 else weighted_prices(total_row)[0]
    charge_price = params.charge_price_override if params.charge_price_override > 0 else weighted_prices(total_row)[1]
    run_days = safe_float(total_row.get("折算天数"))
    payback = payback_summary(
        run_days,
        discharge_price,
        charge_price,
        total_rated,
        total_actual,
        unit_count,
        spec.mode,
        spec.device_cost_wan,
        spec.construction_cost_wan,
        params,
    )
    result = ConfigResult(
        rank=0,
        model=spec.model,
        mode=spec.mode,
        unit_count=unit_count,
        system_power_kw=total_power,
        rated_kwh=total_rated,
        actual_kwh=total_actual,
        run_days=run_days,
        payback_years=payback["payback_years"],
        total_discharge_kwh=safe_float(total_row.get("总放电量(度)")),
        total_charge_kwh=safe_float(total_row.get("总充电量(度)")),
        discharge_price=discharge_price,
        charge_price=charge_price,
        price_spread=discharge_price - charge_price,
        first_year_income_wan=payback["first_year_income_wan"],
        initial_invest_wan=payback["initial_invest_wan"],
        final_cash_flow_wan=payback["final_cash_flow_wan"],
        balance_text=str(total_row.get("综合充放电文本", "")),
        score=0.0,
    )
    return result, monthly_df, payback


def find_top_configs(
    pivot_df: pd.DataFrame,
    specs: list[DeviceSpec],
    unit_range: tuple[int, int],
    run_days_range: tuple[float, float],
    payback_range: tuple[float, float],
    params: ModelParams,
    selected_models: list[str] | None = None,
    selected_modes: list[int] | None = None,
    limit: int = 10,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    results: list[ConfigResult] = []
    detail: dict[str, Any] = {}
    min_units, max_units = unit_range
    for spec in specs:
        if selected_models and spec.model not in selected_models:
            continue
        if selected_modes and spec.mode not in selected_modes:
            continue
        for units in range(min_units, max_units + 1):
            result, monthly_df, payback = evaluate_config(pivot_df, spec, units, params)
            payback_years = result.payback_years
            if payback_years is None:
                continue
            if not (run_days_range[0] <= result.run_days <= run_days_range[1]):
                continue
            if not (payback_range[0] <= payback_years <= payback_range[1]):
                continue
            result.score = payback_years * 1000 - result.final_cash_flow_wan - result.run_days
            key = f"{result.model}-{result.mode}-{result.unit_count}"
            detail[key] = {"monthly": monthly_df, "payback": payback, "result": result}
            results.append(result)
    results.sort(key=lambda r: (r.payback_years if r.payback_years is not None else 999, -r.final_cash_flow_wan, -r.run_days))
    for idx, item in enumerate(results[:limit], start=1):
        item.rank = idx
    df = pd.DataFrame([r.to_dict() for r in results[:limit]])
    if not df.empty:
        df = df.rename(
            columns={
                "rank": "排名",
                "model": "设备型号",
                "mode": "充放模式",
                "unit_count": "台数",
                "system_power_kw": "系统功率(kW)",
                "rated_kwh": "额定容量(kWh)",
                "actual_kwh": "实际容量(kWh)",
                "run_days": "折算运行天数",
                "payback_years": "静态回收期(年)",
                "total_discharge_kwh": "年放电量(kWh)",
                "total_charge_kwh": "年充电量(kWh)",
                "discharge_price": "放电均价(元/kWh)",
                "charge_price": "充电均价(元/kWh)",
                "price_spread": "价差(元/kWh)",
                "first_year_income_wan": "首年电费收入(万元)",
                "initial_invest_wan": "初始投资(万元)",
                "final_cash_flow_wan": "16年累计现金流(万元)",
                "balance_text": "充放状态",
            }
        )
        keep = [col for col in df.columns if col != "score"]
        df = df[keep]
    return df, detail
